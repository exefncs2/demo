# -*- coding: utf-8 -*-

import openpyxl
import time

DATES = time.strftime('%Y%m%d')

class xlsx_builder:
    def __init__(self, file_name: str):
        self.output_file_name = f"{file_name}_{DATES}.xlsx"
        
    def build_blank_xlsx(self, path: str,sheet_name: str):
        file_path = f"{path}/{self.output_file_name}"
        
        # 創建空白檔案
        wb = openpyxl.Workbook()
        
        # set sheet
        ws = wb.active
        
        # 重命名
        ws.title = f"{sheet_name}"
        
        # 寬度
        ws.column_dimensions['A'].width = 15.0
        ws.column_dimensions['B'].width = 15.0
        
        # 標題
        ws['A1'].value = 'test1'
        ws['B1'].value = 'test2'
        
        # 存檔
        wb.save(file_path)
        return file_path
        
    def insert_row(self, file_path: str, sheet_name: str, row: any):
        
        # 開啟檔案
        wb = openpyxl.load_workbook(file_path)
        
        # set sheet
        ws = wb[sheet_name]
        
        # 寫入
        ws.append(row)
        
        # 存檔
        wb.save(file_path)
        